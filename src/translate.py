import os
import sys
import csv
from openpyxl import load_workbook
import argparse


__VERSION__ = "0.0.1"

def main(inputFilePath, outputPath):
    columnNames=["Name","Number","Description","OverrideColor","OverrideStyle","OverrideWeight","OverrideMaterial","ByLevelColor","ByLevelStyle","ByLevelWeight","ByLevelMaterial","DisplayPriority","Transparency","GlobalDisplay","GlobalFreeze","ElementAccess","Plot","OverrideStyleScale","OverrideStyleOriginWidth","OverrideStyleEndWidth","ByLevelStyleScale","ByLevelStyleOriginWidth","ByLevelStyleEndWidth"]
    columnDefault=["", 0, "", 0, 0, 0, None, 0, 0, 0, None, 0, 0.000000, 1, 0, 0, 1, "", "", "", "", "", ""]
    columnMapping=[
        { "ms_nr": 0, "ms_name": "Name", "name": "", "nr": 1 },
        { "ms_nr": 7,  "ms_name": "ByLevelColor", "name": "Farbe", "nr": 2 },
        { "ms_nr": 8,  "ms_name": "ByLevelStyle", "name": "Strichart", "nr": 4 },
        { "ms_nr": 9, "ms_name": "ByLevelWeight", "name": "Strichstaerke", "nr": 3 },
        { "ms_nr": 11, "ms_name": "DisplayPriority", "name": "Prioritaet", "nr": 5 },
        { "ms_nr": 12, "ms_name": "Transparency", "name": "Transparenz", "nr": 6 }
    ]

    xlsx = load_workbook(inputFilePath)
    for sheet in xlsx:
        print(f"processing worksheet \"{sheet.title}\" ...")
        with open(f"{outputPath}\\{sheet.title}.csv", mode='w', newline='') as outfile:
            output = csv.writer(outfile, delimiter=',')

            output.writerow(["%SECTION","Levels"])
            output.writerow("")
            output.writerow(columnNames)

            lastPrefix = ""
            for row in sheet.iter_rows(min_row=3):
                if row[0].value:
                    lastPrefix = row[0].value
                newrow = columnDefault
                for m in columnMapping:
                    val = ""
                    if row[m["nr"]].value:
                        val = row[m["nr"]].value

                    if m["ms_nr"] == 0:
                        val = f"{lastPrefix}{val}"

                    newrow[m["ms_nr"]] = val
                output.writerow(newrow)

            output.writerow("")
            output.writerow(["%SECTION","level-filter"])
            output.writerow("")
            output.writerow(["filter-name","filter-parent","level-group","filter-compose","level-name","level-code","level-description","level-color","level-style","level-weight","level-material","level-element-color","level-element-style","level-element-weight","level-element-material","level-display","level-frozen","level-priority","level-plot","level-used","level-transparency","level-element-access"])

parser = argparse.ArgumentParser()
parser.add_argument("--output-dir", help="Ausgabeverzeichniss", default=os.getcwd())
parser.add_argument("--input", help="Ebenendefinitions-Datei (*.xlsx)")
parser.add_argument("--version", help="Version anzeigen", action="store_true")
args = parser.parse_args()

if args.version:
    print(f"version: {__VERSION__}")
    sys.exit(0)

if not args.input:
    print("Bitte input Datei angeben")
    sys.exit(0)

main(args.input, args.output_dir)
